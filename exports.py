"""
Excel export helpers.

Kept in their own module, separate from app.py's routes, for the same
reason ai_insights.py is: "build a workbook from data" is a distinct
concern from "handle an HTTP request", and keeping them apart makes each
easier to read and test in isolation — every function here takes plain
Python values in and returns an in-memory file, no Flask objects
involved anywhere.

Uses openpyxl directly rather than pandas. pandas is a much heavier
dependency (and pulls in numpy) just to write a handful of columns to a
.xlsx file — openpyxl alone is enough for straightforward tabular
exports like these.
"""

from io import BytesIO 

from openpyxl import Workbook 
from openpyxl .styles import Font ,Alignment ,PatternFill 
from openpyxl .utils import get_column_letter 

HEADER_FONT =Font (name ="Calibri",bold =True ,color ="FFFFFF")
HEADER_FILL =PatternFill (start_color ="2563EB",end_color ="2563EB",fill_type ="solid")
BODY_FONT =Font (name ="Calibri")


def _style_and_size (ws ,column_widths :dict [int ,int ]|None =None ):
    """
    Bolds/colors the header row and sets reasonable column widths, so an
    exported file looks presentable the moment it's opened rather than
    needing a manual "select all, autofit" first. column_widths lets a
    caller override specific columns (1-indexed) that need to be wider
    or narrower than the header-length default.
    """
    column_widths =column_widths or {}
    for cell in ws [1 ]:
        cell .font =HEADER_FONT 
        cell .fill =HEADER_FILL 
        cell .alignment =Alignment (horizontal ="center")

    for row in ws .iter_rows (min_row =2 ):
        for cell in row :
            cell .font =BODY_FONT 

    for idx ,header_cell in enumerate (ws [1 ],start =1 ):
        letter =get_column_letter (idx )
        ws .column_dimensions [letter ].width =column_widths .get (idx ,max (12 ,len (str (header_cell .value ))+4 ))

    ws .freeze_panes ="A2"


def _to_buffer (wb :Workbook )->BytesIO :
    buffer =BytesIO ()
    wb .save (buffer )
    buffer .seek (0 )
    return buffer 


def build_products_workbook (products )->BytesIO :
    """products: a list of Product model instances (already queried/
    ordered by the caller — this function has no database access)."""
    wb =Workbook ()
    ws =wb .active 
    ws .title ="Inventory"

    ws .append (["SKU","Product","Cost","Shipping","Sell Price","Quantity"])
    for p in products :
        ws .append ([p .sku ,p .name ,p .cost_price ,p .shipping_cost ,p .sell_price ,p .quantity ])

    for row in ws .iter_rows (min_row =2 ,min_col =3 ,max_col =5 ):
        for cell in row :
            cell .number_format ="$#,##0.00"

    _style_and_size (ws ,column_widths ={2 :28 })
    return _to_buffer (wb )


def build_sales_workbook (sales ,show_sold_by :bool )->BytesIO :
    """
    sales: a list of Sale model instances, already scoped by the caller
    (this function does no role-based filtering itself — see the
    /sales/export route, which passes only the sales a given user is
    allowed to see, the same way the sales history page already does).
    show_sold_by: whether to include a "Sold by" column — owners get it,
    managers don't, mirroring sales.html's own conditional column.
    """
    wb =Workbook ()
    ws =wb .active 
    ws .title ="Sales"

    headers =["Sale #","Date"]
    if show_sold_by :
        headers .append ("Sold by")
    headers +=["Items","Subtotal","Discount %","Discount Amount","Total"]
    ws .append (headers )

    money_cols =[]
    for s in sales :
        row =[s .id ,s .created_at .strftime ("%Y-%m-%d %H:%M")if s .created_at else ""]
        if show_sold_by :
            row .append (s .user .username if s .user else "")
        row +=[
        len (s .items ),
        s .subtotal ,
        (s .discount_percent /100 )if s .discount_percent else 0 ,
        s .discount_amount if s .discount_percent else 0 ,
        s .total ,
        ]
        ws .append (row )



    subtotal_col =headers .index ("Subtotal")+1 
    discount_pct_col =headers .index ("Discount %")+1 
    discount_amt_col =headers .index ("Discount Amount")+1 
    total_col =headers .index ("Total")+1 

    for row in ws .iter_rows (min_row =2 ):
        row [subtotal_col -1 ].number_format ="$#,##0.00"
        row [discount_pct_col -1 ].number_format ="0.0%"
        row [discount_amt_col -1 ].number_format ="$#,##0.00"
        row [total_col -1 ].number_format ="$#,##0.00"

    _style_and_size (ws )
    return _to_buffer (wb )


def build_employees_workbook (employees ,earnings :dict )->BytesIO :
    """
    employees: a list of User model instances
    earnings: dict keyed by user id -> {"sales_count", "revenue",
    "commission_earned"} — the exact same dict app.py's employees()
    route already builds for the page itself, just reused here.
    """
    wb =Workbook ()
    ws =wb .active 
    ws .title ="Employees"

    ws .append ([
    "Username","Role","Status","Fixed Wage","Commission %",
    "Sales Made","Revenue Generated","Commission Earned",
    ])

    for emp in employees :
        emp_earnings =earnings .get (emp .id )
        ws .append ([
        emp .username ,
        "Shop owner"if emp .role =="owner"else "Sales manager",
        "Active"if emp .active else "Fired",
        emp .fixed_wage ,
        emp .commission_percent /100 ,
        emp_earnings ["sales_count"]if emp_earnings else None ,
        emp_earnings ["revenue"]if emp_earnings else None ,
        emp_earnings ["commission_earned"]if emp_earnings else None ,
        ])

    for row in ws .iter_rows (min_row =2 ):
        row [3 ].number_format ="$#,##0.00"
        row [4 ].number_format ="0.0%"
        row [6 ].number_format ="$#,##0.00"
        row [7 ].number_format ="$#,##0.00"

    _style_and_size (ws )
    return _to_buffer (wb )
